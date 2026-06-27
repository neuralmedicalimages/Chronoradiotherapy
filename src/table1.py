from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from analysis_utils import ensure_dirs
from config import MAJOR_CANCERS, SYNTHETIC_DATA, TABLE_DIR, TIMING_ORDER


def count_percent(series: pd.Series, total: int) -> str:
    count = int(series.sum())
    pct = 100 * count / total if total else np.nan
    return f"{count} ({pct:.1f})"


def median_iqr(series: pd.Series) -> str:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return ""
    q1, q3 = values.quantile([0.25, 0.75])
    return f"{values.median():.1f} ({q1:.1f}-{q3:.1f})"


def add_section(rows: list[dict], label: str) -> None:
    rows.append({"Characteristic": label, **{group: "" for group in ["Overall", *TIMING_ORDER]}})


def summarize_table(df: pd.DataFrame, title: str) -> pd.DataFrame:
    rows: list[dict] = []
    groups = {"Overall": df}
    groups.update({group: df[df["timing_group"] == group] for group in TIMING_ORDER})

    rows.append({"Characteristic": "N", **{name: str(len(sub)) for name, sub in groups.items()}})
    rows.append({"Characteristic": "Age at radiotherapy, median (IQR)", **{name: median_iqr(sub["age_at_rt"]) for name, sub in groups.items()}})

    for variable, label in [
        ("sex", "Sex"),
        ("year_bin", "Year of treatment"),
        ("stage", "Tumor stage"),
        ("kps_group", "KPS group"),
        ("concurrent_chemo", "Concurrent chemotherapy"),
        ("total_dose_group", "Total dose group"),
        ("fraction_dose_group", "Dose per fraction group"),
    ]:
        add_section(rows, label)
        for level in df[variable].dropna().unique():
            row = {"Characteristic": f"  {level}"}
            for name, sub in groups.items():
                row[name] = count_percent(sub[variable].eq(level), len(sub))
            rows.append(row)

    out = pd.DataFrame(rows)
    out.insert(0, "Table", title)
    return out


def write_formatted_excel(result: pd.DataFrame, path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1"
    headers = list(result.columns)
    ws.append(headers)
    for row in result.itertuples(index=False):
        ws.append(list(row))

    thin = Side(style="thin", color="C9D3DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DCEAF5")
    section_fill = PatternFill("solid", fgColor="F2F6FA")
    header_font = Font(name="Arial", size=10, bold=True)
    body_font = Font(name="Arial", size=9)
    section_font = Font(name="Arial", size=9, bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        characteristic = str(ws.cell(row_idx, 2).value or "")
        is_section = characteristic and not characteristic.startswith("  ") and all(
            (ws.cell(row_idx, col_idx).value in [None, ""]) for col_idx in range(3, ws.max_column + 1)
        )
        if is_section:
            for col_idx in range(2, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = section_fill
                ws.cell(row_idx, col_idx).font = section_font
        if characteristic.startswith("  "):
            ws.cell(row_idx, 2).alignment = Alignment(indent=1, vertical="center")

    widths = [18, 34, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    ws.freeze_panes = "C2"
    wb.save(path)


def main() -> None:
    ensure_dirs(TABLE_DIR)
    df = pd.read_csv(SYNTHETIC_DATA)
    tables = [summarize_table(df, "Overall")]
    for cancer in MAJOR_CANCERS:
        tables.append(summarize_table(df[df["cancer_type"] == cancer], cancer))
    result = pd.concat(tables, ignore_index=True)
    out_path = TABLE_DIR / "table1_baseline_characteristics.csv"
    result.to_csv(out_path, index=False)
    excel_path = TABLE_DIR / "table1_baseline_characteristics_formatted.xlsx"
    write_formatted_excel(result, excel_path)
    print(f"Saved baseline characteristics: {out_path}")
    print(f"Saved formatted baseline characteristics: {excel_path}")


if __name__ == "__main__":
    main()
